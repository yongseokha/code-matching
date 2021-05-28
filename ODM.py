from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import copy


lwb_lg = load_workbook("EPCD14CB1B.VP10_working_assy.xlsx")
lws_lg = lwb_lg.active

for row in lws_lg.iter_rows(max_col=2):
    if row[0].value == "품목번호  :":
        model_name = row[1].value
        break

first_name,second_name = model_name.split(".")

lwb_partlist = load_workbook("파트리스트양식참조.xlsx")
lws_partlist = lwb_partlist.active

lws_partlist.title = first_name
lws_partlist["G1"] = first_name

lwb_match = load_workbook("LGIT-SH 자재 코드 매칭.xlsx")
lws_match = lwb_match["매칭표"]

lws_match_sequence = load_workbook("LGIT-SH 자재 코드 매칭.xlsx")
lws_match_sequence = lwb_match["Assy_순서"]

lwb_partlist.create_sheet("매칭표에없는코드")
lws_partlist_unmatch = lwb_partlist["매칭표에없는코드"]

# 틀고정
lws_partlist.freeze_panes = "A4"

row_info = [] # lg code 가 들어있는 row 정보를 row_info 리스트 모두 담기
lg_cnt = 0
for row in lws_lg.iter_rows():
    if row[0].value not in [
        None,
        "Assy별 전개",
        "품목번호  :",
        "BOM이관 :",
        "LEVEL     :",
        "레벨"
    ]:
        row = list(row)
        row.append(lg_cnt)
        lg_cnt += 1
        row_info.append(row)

# assay 분류 담기
packing = []
final = []
manual_insert = []
smd = []
auto_insert = []

# 히트싱크 카운트
cnt_hs = 0
cnt = 0

# 히트싱크 문자열로 담아 놓기
all_assay_str = [
    "PACKING ASS'Y",
    "FINAL ASS'Y",
    "MANUAL INSERT ASS'Y",
    "SMD ASS'Y",
    "AUTO INSERT ASS'Y"
]

# assay 분류하기
for row in row_info:
    if row[0].value == "Assy번호 :" and " HS " in row[4].value:
        cnt_hs += 1
    elif row[0].value == "..2" and row[20].value != None:
        packing.append(row)
    elif row[0].value == "...3":
        if "COATING" in row[5].value or "ADHESIVE" in row[5].value or "Ribbon" in row[5].value or "PACKING" in row[5].value:
            final.append(row)
        else:
            manual_insert.append(row)
    elif row[0].value == "....4":
        if row[3].value == "":
            smd.append(row)
        else:
            # 히트 싱크 갯수에 맞게 동적 변수 생성
            try:
                globals()["HEATSINK{}".format(cnt_hs)].append(row)
            except:
                globals()["HEATSINK{}".format(cnt_hs)] = []
                globals()["HEATSINK{}".format(cnt_hs)].append(row)

    elif row[0].value == ".....5" or row[0].value == "......6" or row[0].value == ".......7":
        auto_insert.append(row)

######### 코드매칭안된 코드 시트 별도 생성###########

row_info_unset = [] # 매칭리스트에 없어 매칭 안된 코드들
row_info_m_all_codename = []
for row in lws_match.iter_rows(min_row=2):
    row_info_m_all_codename.append(row[0].value)

for row in row_info:
    if row[2].value != None and row[2].value not in row_info_m_all_codename:
        row_info_unset.append(list(row))

unmatch_title = [
    "레벨",
    "항번",
    "품번",
    "HSF Flag",
    "Rev",
    "품명",
    "BOM이관",
    "단위",
    "소요량"
]
lws_partlist_unmatch.append(unmatch_title)

for row in row_info_unset:
    lis = []
    for i in range(9):
        lis.append(row[i].value)
    lws_partlist_unmatch.append(lis)

lws_partlist_unmatch.column_dimensions["C"].width = 30
lws_partlist_unmatch.column_dimensions["F"].width = 45

###############


# assay 로 분류한 리스트들 전부 담기
all_assay = [
    packing,
    final,
    manual_insert,
    smd,
    auto_insert
]

# 생성된 히트싱크들 all_assay 담기
for i in range(1,cnt_hs+1)[::-1]:
    all_assay_str.insert(3, "HEATSINK{} ASS'Y".format(i))
    all_assay.insert(3, globals()["HEATSINK{}".format(i)])

# for row in HEATSINK1:
#     for cell in row:
#         print(cell.value, end=" ")
#     print()
#


###  성호 코드에 lg 코드 붙이기 ####
row_info_m = []
for row in lws_match.iter_rows(min_row=2):
    row = list(row)
    if row[0].value != None:
        row_info_m.append(row)
        ro0 = row[0]
        ro1 = row[1]
        ro7 = row[7]
    else:
        row[0] = ro0
        row[1] = ro1
        row[7] = ro7
        row_info_m.append(row)




## 2차원 동적 리스트 만들어 복사 해놓고 맞춰서 담기 ##
all_assay_re = copy.deepcopy(all_assay)

for row in all_assay_re:
    row.clear()

sequence = []

cnt_sequence = 0
for row in lws_match_sequence.iter_cols(min_row=2):
    dic = {}
    for cell in row:
        dic[cell.value] = cnt_sequence
        cnt_sequence += 1
    sequence.append(dic)
    dic = {}

for i in range(cnt_hs):
    sequence.insert(3, {})

for idx,assay in enumerate(all_assay):
    for row in assay:
        for row_m in row_info_m:
            if row[2].value == row_m[0].value:
                row_m = list(row_m)
                row_m.append(row[8])
                row_m.append(row[10])
                row_m.append(row[1])
                row_m.append(row[24])
                if row_m[7].value in sequence[idx].keys():
                    row_m.append(sequence[idx][str(row_m[7].value)])
                else:
                    row_m.append(10000)
                # [lgceode,자재내역,erp,품목명,규격,단위,maker,part,소요량,위치,항번,lg code 순서,순서]
                all_assay_re[idx].append(row_m)

for idx,assay in enumerate(all_assay_re):
    all_assay_re[idx] = sorted(all_assay_re[idx], key=lambda x: (x[12],x[11]))

## 모아놓은 assay 들 partlist 에 뿌리기 ##
cnt_insert = 0
alt_cnt = 1
cnt_insert = []
start = 4
ro0 = ""
for idx, assay in enumerate(all_assay_re):
    no_cnt = 0
    cnt_insert.append(start - 4)
    for row in assay:
        if idx == 0: # -> qty 값 변경
            if row[0].value == ro0:
                if alt_cnt == 1: # alt 1개 첫번째 값 적용
                    if row[10].value != "*R*":

                        lws_partlist["A" + str(start)] = row[0].value
                        lws_partlist["B" + str(start)] = row[1].value
                        lws_partlist["C" + str(start)] = row[2].value

                        lws_partlist["D" + str(start)] = no_cnt
                        lws_partlist["E" + str(start)] = ""
    # [0:lgceode,1:자재내역,2:erp,3:품목명,4:규격,5:단위,6:maker,7:part,8:소요량,9:위치]
                        lws_partlist["F" + str(start)] = ro7
                        lws_partlist["G" + str(start)] = row[3].value
                        lws_partlist["H" + str(start)] = row[4].value
                        lws_partlist["I" + str(start)] = ro5

                        qty = str(float((float(ro8) * 16)))
                        qty_int, qty_float = qty.split(".")
                        qty_int = int(qty_int)
                        qty_float = float(qty_float)
                        qty = float(qty)
                        denominator = 16
                        if qty < 1:
                            lws_partlist["J" + str(start)] = str(qty) + "/" + str(denominator)
                        elif qty >= 1 and qty < 16:
                            if qty_float == 0:
                                lws_partlist["J" + str(start)] = str(int(qty)) + "/" + str(denominator)
                            elif qty_float != 0:
                                lws_partlist["J" + str(start)] = str(qty) + "/" + str(denominator)
                        elif qty >= 16:
                            if qty_float == 0:
                                lws_partlist["J" + str(start)] = str(int(qty / 16))
                            elif qty_float != 0:
                                lws_partlist["J" + str(start)] = str(qty / 16)


                        # qty = float(ro8) * 16
                        # denominator = 16
                        #
                        # if qty < 1 :
                        #     lws_partlist["J" + str(start)] = str(qty) + "/" + str(denominator)
                        # elif qty >= 1 and qty < 16:
                        #     if qty % 16 == 0:
                        #         lws_partlist["J" + str(start)] = str(int(qty)) + "/" + str(denominator)
                        #     elif qty % 16 != 0:
                        #         lws_partlist["J" + str(start)] = str(qty) + "/" + str(denominator)
                        # elif qty >= 16:
                        #     if qty % 16 == 0:
                        #         lws_partlist["J" + str(start)] = str(int(qty)/16)
                        #     elif qty % 16 != 0:
                        #         lws_partlist["J" + str(start)] = str(qty/16)

                        lws_partlist["K" + str(start)] = row[6].value
                        lws_partlist["L" + str(start)] = ro9
                        alt_cnt += 1
                        start += 1
                        ro0 = row[0].value
                    else:
                        lws_partlist["A" + str(start)] = row[0].value
                        lws_partlist["B" + str(start)] = row[1].value
                        lws_partlist["C" + str(start)] = row[2].value

                        lws_partlist["D" + str(start)] = ""
                        lws_partlist["E" + str(start)] = "ALT"
                        # [0:lgceode,1:자재내역,2:erp,3:품목명,4:규격,5:단위,6:maker,7:part,8:소요량,9:위치]
                        lws_partlist["F" + str(start)] = None
                        lws_partlist["G" + str(start)] = row[3].value
                        lws_partlist["H" + str(start)] = row[4].value
                        lws_partlist["I" + str(start)] = None
                        lws_partlist["J" + str(start)] = None


                        lws_partlist["K" + str(start)] = row[6].value
                        lws_partlist["L" + str(start)] = ro9
                        alt_cnt += 1
                        start += 1
                        ro0 = row[0].value

                else: # -> alt 1개초과, 2개 부터 적용
                    lws_partlist["A" + str(start)] = None
                    lws_partlist["B" + str(start)] = None
                    lws_partlist["C" + str(start)] = row[2].value
                    lws_partlist["D" + str(start)] = None
                    lws_partlist["E" + str(start)] = "ALT"
                    lws_partlist["F" + str(start)] = None
                    lws_partlist["G" + str(start)] = row[3].value
                    lws_partlist["H" + str(start)] = row[4].value
                    lws_partlist["I" + str(start)] = None
                    lws_partlist["J" + str(start)] = None
                    lws_partlist["K" + str(start)] = row[6].value
                    lws_partlist["L" + str(start)] = None
                    start += 1

        # 안살려도되고(A B) C(버려)D(cnt) E(alt) F G(버려) H(버려) I J K(버려) L


############ row[0] != ro1 -> 처음 lg 코드 포함한 값 저장만 하고 넘김
            else:
    # [0:lgceode,1:자재내역,2:erp,3:품목명,4:규격,5:단위,6:maker,7:part,8:소요량,9:위치]

                ro0 = row[0].value

                ro7 = row[7].value # part
                ro5 = row[5].value # 규격
                ro8 = row[8].value # qty
                ro9 = row[9].value # location 벨류
                alt_cnt = 1
                if row[10].value != "*R*":
                    no_cnt += 1

######### idx != 0 ################## => qty 값 변경
        else:
            if row[0].value == ro0:
                if alt_cnt == 1:
                    if row[10].value != "*R*":

                        lws_partlist["A" + str(start)] = row[0].value
                        lws_partlist["B" + str(start)] = row[1].value
                        lws_partlist["C" + str(start)] = row[2].value

                        lws_partlist["D" + str(start)] = no_cnt
                        lws_partlist["E" + str(start)] = ""
                        # [0:lgceode,1:자재내역,2:erp,3:품목명,4:규격,5:단위,6:maker,7:part,8:소요량,9:위치]
                        lws_partlist["F" + str(start)] = ro7
                        lws_partlist["G" + str(start)] = row[3].value
                        lws_partlist["H" + str(start)] = row[4].value
                        lws_partlist["I" + str(start)] = ro5
                        lws_partlist["J" + str(start)] = ro8
                        lws_partlist["K" + str(start)] = row[6].value
                        lws_partlist["L" + str(start)] = ro9
                        alt_cnt += 1
                        start += 1
                        ro0 = row[0].value
                    else:
                        lws_partlist["A" + str(start)] = row[0].value
                        lws_partlist["B" + str(start)] = row[1].value
                        lws_partlist["C" + str(start)] = row[2].value

                        lws_partlist["D" + str(start)] = ""
                        lws_partlist["E" + str(start)] = "ALT"
                        # [0:lgceode,1:자재내역,2:erp,3:품목명,4:규격,5:단위,6:maker,7:part,8:소요량,9:위치]
                        lws_partlist["F" + str(start)] = None
                        lws_partlist["G" + str(start)] = row[3].value
                        lws_partlist["H" + str(start)] = row[4].value
                        lws_partlist["I" + str(start)] = None
                        lws_partlist["J" + str(start)] = None
                        lws_partlist["K" + str(start)] = row[6].value
                        lws_partlist["L" + str(start)] = ro9
                        alt_cnt += 1
                        start += 1
                        ro0 = row[0].value
                else:
                    lws_partlist["A" + str(start)] = None
                    lws_partlist["B" + str(start)] = None
                    lws_partlist["C" + str(start)] = row[2].value
                    lws_partlist["D" + str(start)] = None
                    lws_partlist["E" + str(start)] = "ALT"
                    lws_partlist["F" + str(start)] = None
                    lws_partlist["G" + str(start)] = row[3].value
                    lws_partlist["H" + str(start)] = row[4].value
                    lws_partlist["I" + str(start)] = None
                    lws_partlist["J" + str(start)] = None
                    lws_partlist["K" + str(start)] = row[6].value
                    lws_partlist["L" + str(start)] = None
                    start += 1
            # 안살려도되고(A B) C(버려)D(cnt) E(alt) F G(버려) H(버려) I J K(버려) L
            else:
                # [0:lgceode,1:자재내역,2:erp,3:품목명,4:규격,5:단위,6:maker,7:part,8:소요량,9:위치]
                ro0 = row[0].value

                ro7 = row[7].value  # part
                ro5 = row[5].value  # 규격
                ro8 = row[8].value  # qty
                ro9 = row[9].value  # location 벨류
                alt_cnt = 1
                if row[10].value != "*R*":
                    no_cnt += 1


## 처음에 assay 문자열 각 위치에 맞게 뿌리기 ##
packing_zip = list(zip(cnt_insert,all_assay_str))

change_cnt = 0
for i in packing_zip:
    row_cnt = i[0]+4+change_cnt
    lws_partlist.insert_rows(row_cnt)
    lws_partlist.row_dimensions[row_cnt].height = 19.6  # 1 행 높이 50
    for row in lws_partlist.iter_rows(min_row=row_cnt,max_row=row_cnt):
        for cell in row:
            cell.fill = PatternFill(fgColor="31869B", fill_type="solid")


    lws_partlist["D"+str(row_cnt)].value = i[1]
    lws_partlist["D"+str(row_cnt)].font = Font(bold= True, size=15, color="FFFFFF", name="Arial")

    change_cnt += 1

lws_partlist.delete_rows(start+change_cnt,1000)

lwb_partlist.save(str(model_name) + " ODM 생성.xlsx")

